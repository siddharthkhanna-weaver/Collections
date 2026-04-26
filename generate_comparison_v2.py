import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
COLORS = {
    "dark_navy":   "1B2A4A",
    "navy":        "2F4073",
    "header_bg":   "1B2A4A",
    "cat_bg":      "E8EDF5",
    "cat_font":    "1B2A4A",
    "white":       "FFFFFF",
    "light_gray":  "F7F8FA",
    "mid_gray":    "E9ECF0",
    "border":      "C5CDD8",
    "yes_bg":      "E6F4EA",
    "yes_font":    "1E7D34",
    "no_bg":       "FDE8E8",
    "no_font":     "C53030",
    "partial_bg":  "FFF8E1",
    "partial_font":"B7791F",
    "accent1":     "3B82F6",  # Omnifin blue
    "accent2":     "8B5CF6",  # Spocto purple
    "accent3":     "10B981",  # ENCollect green
    "score_high":  "22C55E",
    "score_mid":   "F59E0B",
    "score_low":   "EF4444",
    "title_bg":    "0F172A",
    "subtitle":    "64748B",
}

thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

header_font = Font(name="Calibri", bold=True, size=11, color=COLORS["white"])
header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cat_font = Font(name="Calibri", bold=True, size=11, color=COLORS["cat_font"])
cat_fill = PatternFill(start_color=COLORS["cat_bg"], end_color=COLORS["cat_bg"], fill_type="solid")

feature_font = Font(name="Calibri", size=10.5, color="334155")
feature_font_bold = Font(name="Calibri", size=10.5, color="334155", bold=True)
wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

yes_fill = PatternFill(start_color=COLORS["yes_bg"], end_color=COLORS["yes_bg"], fill_type="solid")
yes_font = Font(name="Calibri", bold=True, size=10, color=COLORS["yes_font"])
no_fill = PatternFill(start_color=COLORS["no_bg"], end_color=COLORS["no_bg"], fill_type="solid")
no_font = Font(name="Calibri", bold=True, size=10, color=COLORS["no_font"])
partial_fill = PatternFill(start_color=COLORS["partial_bg"], end_color=COLORS["partial_bg"], fill_type="solid")
partial_font = Font(name="Calibri", size=10, color=COLORS["partial_font"])

row1_fill = PatternFill(start_color=COLORS["white"], end_color=COLORS["white"], fill_type="solid")
row2_fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")

# ─────────────────────────────────────────────
# DATA — Structured with categories & improved
# ─────────────────────────────────────────────
# Format: (feature, omnifin, spocto, encollect, importance)
# importance: "High" / "Medium" / "Low"

categories = {
    "Integration & Data Flow": [
        ("Case Information Flow to Collection Module", "LMS and Collection integrated natively", "Integration via API or File Transfer", "Integration via API or File Transfer", "High"),
        ("LOS / LMS Data Sync (Loan Details, EMI, Outstanding)", "Native sync — real-time", "API-based sync — near real-time", "API-based sync — near real-time", "High"),
        ("Payment Confirmation Mechanism", "Manual entry by RM / Collection Officer", "API-based payment confirmation", "API-based payment confirmation", "High"),
    ],
    "Communication & Orchestration": [
        ("Multi-channel Communication (SMS, Email, WhatsApp)", "No", "Yes", "Yes", "High"),
        ("Communication to Customer on Dispositions (PTP, NC, Door Locked, etc.)", "No", "Yes", "Yes", "Medium"),
        ("Cash Receipt Communication to Customer", "By uploading Pay-in Slip", "Yes — Automated", "Yes — Automated", "Medium"),
    ],
    "Allocation & Strategy": [
        ("Concurrent Allocation to Field, Call Centre & Digital", "No", "Yes", "Yes", "High"),
        ("Cross-Channel Feedback as Trail of Events per Case", "No", "Yes", "Yes", "High"),
        ("Field Allocation Based on RM / Officer Base Location", "Yes", "Yes", "Yes", "Medium"),
        ("Alternate Contact Information Repository", "No", "Yes", "Yes", "Medium"),
    ],
    "Call Centre Platform": [
        ("Dedicated Call Centre Platform", "No", "Yes", "Yes", "High"),
        ("Dialler Integration (Auto / Predictive / Progressive)", "No", "Yes", "Yes", "High"),
    ],
    "Field Operations": [
        ("Dedicated Field Team App", "Yes", "Yes", "Yes", "High"),
        ("Route Optimization in App", "No", "Yes", "Yes", "Medium"),
        ("Agent Route Tracking (Live / Historical)", "No", "Yes", "Yes", "Medium"),
        ("Photo Upload with Geo-tagging", "Yes", "Yes", "Yes", "Medium"),
        ("Event-based Geotagging", "Yes", "Yes", "Yes", "Medium"),
    ],
    "Disposition & Case Management": [
        ("Disposition Management (Configurable)", "Yes", "Yes", "Yes", "High"),
        ("Probability to Pay — AI/ML Scoring", "No", "Yes", "Yes", "High"),
    ],
    "Digital Payment Collection": [
        ("Payment Link Generation", "Yes", "Yes", "Yes", "High"),
        ("Dynamic QR Code Generation", "Yes", "No", "Yes", "Medium"),
        ("Apportionment of Digital Collection (Manual Entry)", "Yes", "No", "Yes", "Low"),
    ],
}

# ─────────────────────────────────────────────
# SHEET 1 — Detailed Comparison
# ─────────────────────────────────────────────
ws = wb.active
ws.title = "Detailed Comparison"
ws.sheet_properties.tabColor = COLORS["accent1"]

# Column widths
col_widths = {"A": 6, "B": 50, "C": 30, "D": 30, "E": 30, "F": 14}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── Title Row ──
ws.merge_cells("A1:F1")
title_cell = ws.cell(row=1, column=1, value="Collections Platform Comparison — Omnifin vs Spocto vs ENCollect")
title_cell.font = Font(name="Calibri", bold=True, size=16, color=COLORS["white"])
title_cell.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 48

# ── Subtitle Row ──
ws.merge_cells("A2:F2")
sub_cell = ws.cell(row=2, column=1, value="PeopleHome — Debt Collection Module Evaluation  |  Last Updated: March 2026")
sub_cell.font = Font(name="Calibri", size=10, color=COLORS["subtitle"], italic=True)
sub_cell.fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 28

# ── Headers ──
headers = ["#", "Feature", "Omnifin", "Spocto", "ENCollect", "Importance"]
accent_fills = {
    2: PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
    3: PatternFill(start_color=COLORS["accent1"], end_color=COLORS["accent1"], fill_type="solid"),
    4: PatternFill(start_color=COLORS["accent2"], end_color=COLORS["accent2"], fill_type="solid"),
    5: PatternFill(start_color=COLORS["accent3"], end_color=COLORS["accent3"], fill_type="solid"),
}

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = accent_fills.get(col_idx, header_fill)
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 36

# ── Data Rows ──
row_num = 4
feature_num = 0

def style_value_cell(cell, value):
    val_upper = str(value).strip().upper()
    cell.alignment = wrap_center
    cell.border = thin_border
    if val_upper == "YES" or val_upper.startswith("YES"):
        cell.fill = yes_fill
        cell.font = yes_font
        if val_upper == "YES":
            cell.value = "✓  Yes"
        else:
            cell.value = "✓  " + str(value)
    elif val_upper == "NO":
        cell.fill = no_fill
        cell.font = no_font
        cell.value = "✗  No"
    else:
        cell.fill = partial_fill
        cell.font = partial_font

def style_importance(cell, value):
    cell.alignment = wrap_center
    cell.border = thin_border
    cell.font = Font(name="Calibri", bold=True, size=10)
    if value == "High":
        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10, color="166534")
        cell.value = "● High"
    elif value == "Medium":
        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10, color="92400E")
        cell.value = "● Medium"
    else:
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=10, color="6B7280")
        cell.value = "● Low"

for cat_name, features in categories.items():
    # Category row
    ws.merge_cells(f"A{row_num}:F{row_num}")
    cat_cell = ws.cell(row=row_num, column=1, value=f"  {cat_name}")
    cat_cell.font = cat_font
    cat_cell.fill = cat_fill
    cat_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 7):
        ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=c).fill = cat_fill
    ws.row_dimensions[row_num].height = 32
    row_num += 1

    for feat, omnifin, spocto, encollect, importance in features:
        feature_num += 1
        bg = row1_fill if feature_num % 2 == 1 else row2_fill

        # Number
        num_cell = ws.cell(row=row_num, column=1, value=feature_num)
        num_cell.font = Font(name="Calibri", size=10, color=COLORS["subtitle"])
        num_cell.alignment = wrap_center
        num_cell.fill = bg
        num_cell.border = thin_border

        # Feature name
        feat_cell = ws.cell(row=row_num, column=2, value=feat)
        feat_cell.font = feature_font
        feat_cell.alignment = wrap_left
        feat_cell.fill = bg
        feat_cell.border = thin_border

        # Values
        for col_idx, val in [(3, omnifin), (4, spocto), (5, encollect)]:
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            style_value_cell(cell, val)

        # Importance
        imp_cell = ws.cell(row=row_num, column=6)
        style_importance(imp_cell, importance)

        ws.row_dimensions[row_num].height = 38
        row_num += 1

# ── Summary Row ──
row_num += 1
ws.merge_cells(f"A{row_num}:F{row_num}")
ws.cell(row=row_num, column=1).fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
ws.row_dimensions[row_num].height = 6

row_num += 1
ws.merge_cells(f"A{row_num}:B{row_num}")
lbl = ws.cell(row=row_num, column=1, value="Feature Coverage Score")
lbl.font = Font(name="Calibri", bold=True, size=12, color=COLORS["cat_font"])
lbl.alignment = Alignment(horizontal="left", vertical="center")

# Calculate scores
def calc_score(vendor_idx):
    total = 0
    yes_count = 0
    for features in categories.values():
        for feat in features:
            total += 1
            val = str(feat[vendor_idx]).strip().upper()
            if val == "YES" or val.startswith("YES"):
                yes_count += 1
    return yes_count, total

omni_yes, total = calc_score(1)
spocto_yes, _ = calc_score(2)
enc_yes, _ = calc_score(3)

for col_idx, (yes_c, label, color) in [
    (3, (omni_yes, "Omnifin", COLORS["accent1"])),
    (4, (spocto_yes, "Spocto", COLORS["accent2"])),
    (5, (enc_yes, "ENCollect", COLORS["accent3"])),
]:
    pct = round(yes_c * 100 / total)
    cell = ws.cell(row=row_num, column=col_idx, value=f"{yes_c}/{total} ({pct}%)")
    cell.font = Font(name="Calibri", bold=True, size=12, color=color)
    cell.alignment = wrap_center
    cell.border = thin_border
    if pct >= 70:
        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    elif pct >= 50:
        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

ws.row_dimensions[row_num].height = 40

# Freeze and filter
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{row_num - 2}"

# Print settings
ws.print_title_rows = "1:3"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ─────────────────────────────────────────────
# SHEET 2 — Scorecard Summary
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Scorecard")
ws2.sheet_properties.tabColor = COLORS["accent2"]

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 20

# Title
ws2.merge_cells("A1:D1")
t = ws2.cell(row=1, column=1, value="Category-wise Scorecard")
t.font = Font(name="Calibri", bold=True, size=14, color=COLORS["white"])
t.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 42

# Headers
for col_idx, (h, color) in enumerate([
    ("Category", COLORS["header_bg"]),
    ("Omnifin", COLORS["accent1"]),
    ("Spocto", COLORS["accent2"]),
    ("ENCollect", COLORS["accent3"]),
], 1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = header_align
    cell.border = thin_border
ws2.row_dimensions[2].height = 32

# Category scores
row_num = 3
for cat_name, features in categories.items():
    cat_total = len(features)
    scores = {}
    for vendor_idx, vendor in [(1, "Omnifin"), (2, "Spocto"), (3, "ENCollect")]:
        yes_c = sum(1 for feat in features if str(feat[vendor_idx]).strip().upper() == "YES" or str(feat[vendor_idx]).strip().upper().startswith("YES"))
        scores[vendor] = yes_c

    bg = row1_fill if row_num % 2 == 1 else row2_fill

    cell = ws2.cell(row=row_num, column=1, value=cat_name)
    cell.font = Font(name="Calibri", bold=True, size=10.5, color="334155")
    cell.alignment = wrap_left
    cell.fill = bg
    cell.border = thin_border

    for col_idx, vendor in [(2, "Omnifin"), (3, "Spocto"), (4, "ENCollect")]:
        sc = scores[vendor]
        pct = round(sc * 100 / cat_total)
        cell = ws2.cell(row=row_num, column=col_idx, value=f"{sc}/{cat_total} ({pct}%)")
        cell.alignment = wrap_center
        cell.border = thin_border
        if pct >= 75:
            cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            cell.font = Font(name="Calibri", bold=True, size=10.5, color="166534")
        elif pct >= 50:
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            cell.font = Font(name="Calibri", bold=True, size=10.5, color="92400E")
        else:
            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            cell.font = Font(name="Calibri", bold=True, size=10.5, color="991B1B")

    ws2.row_dimensions[row_num].height = 34
    row_num += 1

# Total row
row_num += 1
ws2.merge_cells(f"A{row_num}:D{row_num}")
ws2.cell(row=row_num, column=1).fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
ws2.row_dimensions[row_num].height = 4

row_num += 1
cell = ws2.cell(row=row_num, column=1, value="OVERALL SCORE")
cell.font = Font(name="Calibri", bold=True, size=12, color=COLORS["cat_font"])
cell.alignment = wrap_left
cell.border = thin_border

for col_idx, (yes_c, color) in [(2, (omni_yes, COLORS["accent1"])), (3, (spocto_yes, COLORS["accent2"])), (4, (enc_yes, COLORS["accent3"]))]:
    pct = round(yes_c * 100 / total)
    cell = ws2.cell(row=row_num, column=col_idx, value=f"{yes_c}/{total} ({pct}%)")
    cell.font = Font(name="Calibri", bold=True, size=13, color=color)
    cell.alignment = wrap_center
    cell.border = thin_border
    if pct >= 70:
        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    elif pct >= 50:
        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

ws2.row_dimensions[row_num].height = 42
ws2.freeze_panes = "A3"

# ─────────────────────────────────────────────
# SHEET 3 — Legend & Notes
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Legend & Notes")
ws3.sheet_properties.tabColor = COLORS["accent3"]
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 60

ws3.merge_cells("A1:B1")
t = ws3.cell(row=1, column=1, value="Legend & Reading Guide")
t.font = Font(name="Calibri", bold=True, size=14, color=COLORS["white"])
t.fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

legend_data = [
    ("Color Coding", ""),
    ("✓  Yes (Green)", "Feature is fully available out-of-the-box"),
    ("✗  No (Red)", "Feature is not available"),
    ("Text (Amber)", "Partially available or with caveats — read the detail"),
    ("", ""),
    ("Importance Levels", ""),
    ("● High", "Critical for collections operations — must-have for go-live"),
    ("● Medium", "Important for efficiency — should be available within first quarter"),
    ("● Low", "Nice-to-have — can be addressed in later phases"),
    ("", ""),
    ("Score Thresholds", ""),
    ("≥ 75% (Green)", "Strong coverage — vendor meets most requirements"),
    ("50–74% (Amber)", "Moderate coverage — gaps need to be addressed"),
    ("< 50% (Red)", "Weak coverage — significant gaps, needs custom development or alternative"),
]

for i, (key, val) in enumerate(legend_data, 2):
    k = ws3.cell(row=i, column=1, value=key)
    v = ws3.cell(row=i, column=2, value=val)
    k.font = Font(name="Calibri", bold=True, size=10.5, color="334155")
    v.font = Font(name="Calibri", size=10.5, color="64748B")
    k.alignment = wrap_left
    v.alignment = wrap_left
    if not val and key:  # Section headers
        k.font = Font(name="Calibri", bold=True, size=11, color=COLORS["cat_font"])
        k.fill = cat_fill
        v.fill = cat_fill
    k.border = thin_border
    v.border = thin_border

# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
output = "/Users/siddharthkhanna/Collections/Omnifin Vs Spocto V2.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"\nScores — Omnifin: {omni_yes}/{total} ({round(omni_yes*100/total)}%) | Spocto: {spocto_yes}/{total} ({round(spocto_yes*100/total)}%) | ENCollect: {enc_yes}/{total} ({round(enc_yes*100/total)}%)")
